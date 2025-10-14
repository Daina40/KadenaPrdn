import json
from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404, HttpResponse
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from Style_Information_app.models import Customer, StyleInfo, StyleDescription, Comment, StyleImage
from collections import defaultdict
from django.contrib import messages
from django.core.files.storage import default_storage
import os

def style_info_add(request):
    if request.method == "POST":
        customer_name = request.POST.get("customer_name", "").upper()
        style_no = request.POST.get("style_no", "").upper()
        season = request.POST.get("season", "").upper()
        production_line = request.POST.get("production_line", "").upper()
        order_qty = request.POST.get("order_qty", 0)
        apm = request.POST.get("apm", "").capitalize()
        technician = request.POST.get("technician", "").capitalize()
        qc = request.POST.get("qc", "").capitalize()
        qa = request.POST.get("qa", "").capitalize()
        tqs = request.POST.get("tqs", "").capitalize()

        customer, _ = Customer.objects.get_or_create(customer_name=customer_name)
        style = StyleInfo.objects.create(
            customer=customer,
            style_no=style_no,
            season=season,
            production_line=production_line,
            order_qty=order_qty,
            apm=apm,
            technician=technician,
            qc=qc,
            qa=qa,
            tqs=tqs,
            source='overview' 
        )

        if request.POST.get("style_description"):  
            StyleDescription.objects.create(
                style=style,
                style_description=request.POST.get("style_description")
            )

        return redirect('add_style_overview')

    return render(request, 'style_information/style_info_add.html')

def add_style_overview(request):
    overview_styles = StyleInfo.objects.filter(source='overview').prefetch_related("descriptions", "customer").all()

    detail_styles = StyleInfo.objects.filter(source='detail').prefetch_related("descriptions")

    detail_desc_map = defaultdict(set)
    for d in detail_styles:
        for desc in d.descriptions.all():
            detail_desc_map[d.style_no].add(desc)

    customer_grouped = defaultdict(lambda: {"styles": {}, "rowspan": 0})

    for s in overview_styles:
        cust = s.customer.customer_name
        style_no = s.style_no

        if style_no not in customer_grouped[cust]["styles"]:
            customer_grouped[cust]["styles"][style_no] = {
                "rows": [],
                "descriptions": {},
            }

        row_key = (
            s.season,
            s.production_line,
            s.apm,
            s.technician,
            s.qc,
            s.qa,
            s.tqs,
        )
        if row_key not in [
            (
                r.season,
                r.production_line,
                r.apm,
                r.technician,
                r.qc,
                r.qa,
                r.tqs,
            )
            for r in customer_grouped[cust]["styles"][style_no]["rows"]
        ]:
            customer_grouped[cust]["styles"][style_no]["rows"].append(s)
            customer_grouped[cust]["rowspan"] += 1

        combined_descs = list(s.descriptions.all()) + list(detail_desc_map.get(style_no, []))
        for d in combined_descs:
            if d.style_description not in customer_grouped[cust]["styles"][style_no]["descriptions"]:
                customer_grouped[cust]["styles"][style_no]["descriptions"][d.style_description] = d

    merged_customers = []
    for customer, cdata in customer_grouped.items():
        styles_list = []
        for style_no, sdata in cdata["styles"].items():
            styles_list.append({
                "style_no": style_no,
                "rows": sdata["rows"],
                "descriptions": list(sdata["descriptions"].values()),
                "rowspan": len(sdata["rows"]),
            })
        merged_customers.append({
            "customer": customer,
            "styles": styles_list,
            "rowspan": cdata["rowspan"],
        })

    return render(request, "style_information/add_style_overview.html", {
        "customers": merged_customers
    })

def delete_add_style_overview(request, id):
    style = get_object_or_404(StyleInfo, id=id)
    style.delete()
    return redirect('add_style_overview')

def edit_add_style_overview(request, id):
    style = get_object_or_404(StyleInfo, id=id)

    if request.method == "POST":
        style.customer.customer_name = request.POST.get("customer_name", "").upper()
        style.style_no = request.POST.get("style_no", "").upper()
        style.season = request.POST.get("season", "").upper()
        style.production_line = request.POST.get("production_line", "").upper()
        style.apm = request.POST.get("apm", "").capitalize()
        style.technician = request.POST.get("technician", "").capitalize()
        style.qc = request.POST.get("qc", "").capitalize()
        style.qa = request.POST.get("qa", "").capitalize()
        style.tqs = request.POST.get("tqs", "").capitalize()

        style.customer.save()
        style.save()

        # Update or create description if provided
        style_desc = request.POST.get("style_description", "")
        if style_desc:
            StyleDescription.objects.update_or_create(
                style=style,
                defaults={"style_description": style_desc}
            )

        return redirect("add_style_overview")

    # Pre-fill existing values in context
    context = {
        "edit": True,
        "customer_name": style.customer.customer_name,
        "style_no": style.style_no,
        "season": style.season,
        "production_line": style.production_line,
        "apm": style.apm,
        "technician": style.technician,
        "qc": style.qc,
        "qa": style.qa,
        "tqs": style.tqs,
        "style_id": style.id,
        "style_description": style.descriptions.first().style_description if style.descriptions.exists() else "",
    }

    return render(request, "style_information/style_info_add.html", context)


def style_detail(request, style_id):
    style = get_object_or_404(
        StyleInfo.objects.prefetch_related("images", "descriptions__images", "comments", "customer"),
        id=style_id
    )

    descriptions = style.descriptions.all()

    comments_dict = {
        desc.id: {
            c.process.strip(): c.comment_text
            for c in style.comments.filter(description=desc)
        }
        for desc in descriptions
    }

    styles = StyleInfo.objects.all()
    customers = styles.values_list("customer__customer_name", flat=True).distinct()
    seasons = styles.values_list("season", flat=True).distinct()
    lines = styles.values_list("production_line", flat=True).distinct()
    apms = styles.values_list("apm", flat=True).distinct()
    technicians = styles.values_list("technician", flat=True).distinct()
    qcs = styles.values_list("qc", flat=True).distinct()
    qas = styles.values_list("qa", flat=True).distinct()
    tqss = styles.values_list("tqs", flat=True).distinct()
    style_nos = styles.values_list("style_no", flat=True).distinct()

    context = {
        "style": style,
        "descriptions": descriptions,
        "comments_dict": comments_dict,
        "customers": customers,
        "seasons": seasons,
        "lines": lines,
        "apms": apms,
        "technicians": technicians,
        "qcs": qcs,
        "qas": qas,
        "tqss": tqss,
        "style_nos": style_nos,
        "read_only": False,
        "form_action": "save_style_info",
    }
    return render(request, "style_information/style_detail.html", context)


def save_comments(request, style_id):
    style = get_object_or_404(StyleInfo, id=style_id)

    if request.method == "POST":
        try:
            data = json.loads(request.body.decode("utf-8"))
            process = data.get("process", "").strip()
            comment_text = data.get("comment", "").strip()
            description_id = data.get("description_id")

            if not process or not comment_text or not description_id:
                return JsonResponse({"success": False, "error": "Invalid input."}, status=400)

            description = get_object_or_404(StyleDescription, id=description_id, style=style)

            comment_obj, created = Comment.objects.update_or_create(
                style=style,
                description=description,
                process=process,
                defaults={"comment_text": comment_text}
            )

            return JsonResponse({
                "success": True,
                "process": process,
                "comment": comment_text,
                "description_id": description.id,
                "created": created
            })

        except json.JSONDecodeError:
            return JsonResponse({"success": False, "error": "Invalid JSON."}, status=400)

    return JsonResponse({"success": False, "error": "Invalid request method."}, status=405)

def delete_comment(request, style_id):
    if request.method == "POST":
        style = get_object_or_404(StyleInfo, id=style_id)
        try:
            data = json.loads(request.body)
            process = (data.get("process") or "").strip()
            description_id = data.get("description_id")

            if not process or not description_id:
                return JsonResponse({"success": False, "error": "Invalid data"}, status=400)

            description = get_object_or_404(StyleDescription, id=description_id, style=style)
            comment = style.comments.filter(description=description, process=process).first()

            if comment:
                comment.delete()
                return JsonResponse({"success": True})
            else:
                return JsonResponse({"success": False, "error": "Comment not found"}, status=404)

        except json.JSONDecodeError:
            return JsonResponse({"success": False, "error": "Invalid JSON"}, status=400)

    return JsonResponse({"success": False, "error": "Invalid request method"}, status=405)

def save_style_info(request, style_id):
    original_style = get_object_or_404(StyleInfo, id=style_id)

    if request.method == "POST":
        new_style = StyleInfo.objects.create(
            customer=original_style.customer,
            style_no=original_style.style_no,
            source="detail",
            season=request.POST.get("season"),
            production_line=request.POST.get("production_line"),
            apm=request.POST.get("apm"),
            technician=request.POST.get("technician"),
            qc=request.POST.get("qc"),
            qa=request.POST.get("qa"),
            tqs=request.POST.get("tqs"),
            program=request.POST.get("program"),
        )

        order_qty = request.POST.get("order_qty")
        if not order_qty or not order_qty.isdigit() or int(order_qty) <= 0:
            messages.error(request, "Order quantity is required and must be a positive number.")
            return redirect("style_detail", style_id=style_id)

        new_style.order_qty = int(order_qty)
        new_style.save()

        desc_map = {}

        for desc in original_style.descriptions.all():
            new_desc = StyleDescription.objects.create(
                style=new_style,
                style_description=desc.style_description
            )
            desc_map[desc.id] = new_desc

            for img in desc.images.all():
                if not img.image_url or not img.style_id:
                    continue

                StyleImage.objects.create(
                    style=new_style,
                    description=new_desc,
                    image_name=img.image_name,
                    image_url=img.image_url,
                )

        for comment in original_style.comments.all():
            old_desc = comment.description
            new_desc = desc_map.get(old_desc.id) if old_desc else None

            Comment.objects.create(
                style=new_style,
                description=new_desc,
                responsible_person=comment.responsible_person,
                process=comment.process,
                comment_text=comment.comment_text,
            )

        messages.success(request, "Style information saved successfully.")
        return redirect("style_saved_table")

    return redirect("style_detail", style_id=style_id)

def style_saved_table(request):
    styles = StyleInfo.objects.filter(source='detail').prefetch_related("descriptions", "customer").order_by('-created_at')
    customer_grouped = defaultdict(lambda: {"styles": {}, "rowspan": 0})

    for s in styles:
        print("→", s.id, s.style_no, s.customer, s.source)
        cust_name = s.customer.customer_name if s.customer else "Unknown Customer"
        style_no = s.style_no
        s.first_description = s.descriptions.first()

        if style_no not in customer_grouped[cust_name]["styles"]:
            customer_grouped[cust_name]["styles"][style_no] = {"rows": []}

        customer_grouped[cust_name]["styles"][style_no]["rows"].append(s)
        customer_grouped[cust_name]["rowspan"] += 1

    merged_customers = []
    for customer, cdata in customer_grouped.items():
        styles_list = []
        for style_no, sdata in cdata["styles"].items():
            styles_list.append({
                "style_no": style_no,
                "rows": sdata["rows"],
                "rowspan": len(sdata["rows"]),
            })
        merged_customers.append({
            "customer": customer,
            "styles": styles_list,
            "rowspan": cdata["rowspan"],
        })

    return render(request, "style_information/style_saved_table.html", {
        "customers": merged_customers
    })

def style_saved_table_delete(request, style_id):
    if request.method == "POST":
        style = get_object_or_404(StyleInfo, id=style_id)
        style.delete()
        messages.success(request, f"Style '{style.style_no}' deleted successfully.")
        return redirect('style_saved_table')
    else:
        messages.error(request, "Invalid request method.")
        return redirect('style_saved_table')

def style_saved_table_edit(request, style_id):
    style = get_object_or_404(
        StyleInfo.objects.prefetch_related("images", "descriptions__images", "comments", "customer"),
        id=style_id
    )

    if request.method == "POST":
        style.production_line = request.POST.get("production_line") or style.production_line
        style.technician = request.POST.get("technician") or style.technician
        style.season = request.POST.get("season") or style.season
        style.order_qty = request.POST.get("order_qty") or style.order_qty
        style.qc = request.POST.get("qc") or style.qc
        style.apm = request.POST.get("apm") or style.apm
        style.qa = request.POST.get("qa") or style.qa
        style.tqs = request.POST.get("tqs") or style.tqs
        style.program = request.POST.get("program") or style.program

        style.source = "detail"
        style.save()
        messages.success(request, f"Style '{style.style_no}' updated successfully.")
        return redirect("style_saved_table")

    descriptions = style.descriptions.all()
    comments_dict = {
        desc.id: {
            c.process.strip(): c.comment_text
            for c in style.comments.filter(description=desc)
        }
        for desc in descriptions
    }

    styles = StyleInfo.objects.all()
    customers = styles.values_list("customer__customer_name", flat=True).distinct()
    seasons = styles.values_list("season", flat=True).distinct()
    lines = styles.values_list("production_line", flat=True).distinct()
    apms = styles.values_list("apm", flat=True).distinct()
    technicians = styles.values_list("technician", flat=True).distinct()
    qcs = styles.values_list("qc", flat=True).distinct()
    qas = styles.values_list("qa", flat=True).distinct()
    tqss = styles.values_list("tqs", flat=True).distinct()
    style_nos = styles.values_list("style_no", flat=True).distinct()

    context = {
        "style": style,
        "descriptions": descriptions,
        "comments_dict": comments_dict,
        "customers": customers,
        "seasons": seasons,
        "lines": lines,
        "apms": apms,
        "technicians": technicians,
        "qcs": qcs,
        "qas": qas,
        "tqss": tqss,
        "style_nos": style_nos,
        "read_only": False,
        "form_action": "style_saved_table_edit",
    }

    return render(request, "style_information/style_detail.html", context)

def upload_style_image(request):
    if request.method == "POST" and request.FILES.get("image"):
        style_id = request.POST.get("style_id")
        description_id = request.POST.get("description_id")
        file = request.FILES["image"]

        file_path = default_storage.save(os.path.join("style_images", file.name), file)
        file_url = default_storage.url(file_path)

        style = StyleInfo.objects.get(id=style_id)
        description = StyleDescription.objects.get(id=description_id)

        image = StyleImage.objects.create(
            style=style,
            description=description,
            image_name=file.name,
            image_url=file_url
        )

        return JsonResponse({
            "success": True,
            "image_id": image.id,
            "image_name": image.image_name,
            "image_url": image.image_url,
        })

    return JsonResponse({"success": False, "error": "Invalid request"})

def delete_style_image(request, image_id):
    if request.method == "POST":
        try:
            image = StyleImage.objects.get(id=image_id)

            if image.image_url:
                file_path = image.image_url.replace("/media/", "")
                if default_storage.exists(file_path):
                    default_storage.delete(file_path)

            image.delete()
            return JsonResponse({"success": True})
        except StyleImage.DoesNotExist:
            return JsonResponse({"success": False, "error": "Image not found"})
    return JsonResponse({"success": False, "error": "Invalid request"})

def style_view(request, style_id):
    style = get_object_or_404(
        StyleInfo.objects.prefetch_related(
            "images", "descriptions__images", "comments", "customer"
        ),
        id=style_id
    )

    descriptions = style.descriptions.all()

    comments_dict = {
        desc.id: {
            c.process: c.comment_text
            for c in style.comments.filter(description=desc)
        }
        for desc in descriptions
    }

    styles = StyleInfo.objects.all()
    customers = styles.values_list("customer__customer_name", flat=True).distinct()
    seasons = styles.values_list("season", flat=True).distinct()
    lines = styles.values_list("production_line", flat=True).distinct()
    apms = styles.values_list("apm", flat=True).distinct()
    technicians = styles.values_list("technician", flat=True).distinct()
    qcs = styles.values_list("qc", flat=True).distinct()
    qas = styles.values_list("qa", flat=True).distinct()
    tqss = styles.values_list("tqs", flat=True).distinct()
    style_nos = styles.values_list("style_no", flat=True).distinct()

    context = {
        "style": style,
        "descriptions": descriptions,
        "comments_dict": comments_dict,
        "customers": customers,
        "seasons": seasons,
        "lines": lines,
        "apms": apms,
        "technicians": technicians,
        "qcs": qcs,
        "qas": qas,
        "tqss": tqss,
        "style_nos": style_nos,
        "read_only": True,
    }

    return render(request, "style_information/style_detail.html", context)

def set_border(ws, start_row, start_col, end_row, end_col, border):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = border
            
def download_style_excel(request, style_id):
    style = get_object_or_404(
        StyleInfo.objects.prefetch_related("descriptions__comments", "customer"),
        id=style_id
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary of style information"

    # ---- Styles ----
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # ---- Sheet column widths ----
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    # ---- Header (exact layout like template) ----
    # Merge main title
    ws.merge_cells("A2:F2")
    ws["A2"] = "Summary of style information"
    ws["A2"].font = Font(bold=True, size=14)
    ws["A2"].alignment = center

    # Date (right side)
    ws["E3"] = "Date:"
    ws["F3"] = style.created_at.strftime("%d-%b-%Y") if style.created_at else ""
    ws["E3"].font = bold
    ws["E3"].alignment = right
    ws["F3"].alignment = center

    # ---- Information Section (left and right grouping) ----
    # Left-side info
    info_rows_left = [
        ("Customer Name:", style.customer.customer_name if style.customer else ""),
        ("Season:", style.season or ""),
        ("Style:", style.style_no or ""),
        ("Style Description:", "\n".join(d.style_description for d in style.descriptions.all())),
        ("Program:", style.program or ""),
    ]

    info_rows_middle = [
        ("Production Line:", style.production_line or ""),
        ("Order Qty:", f"{style.order_qty} pcs" if style.order_qty else ""),
        ("APM:", style.apm or ""),
    ]
    
    # Right-side info (Production / Technician side)
    info_rows_right = [
        ("Technician:", style.technician or ""),
        ("QC:", style.qc or ""),
        ("QA:", style.qa or ""),
        ("TQS:", style.tqs or ""),
    ]

    # Start rows at same height as original
    start_row = 4
    left_row = start_row
    middle_row = start_row
    right_row = start_row

    # Write left info (cols A–B)
    for label, value in info_rows_left:
        ws[f"A{left_row}"] = label
        ws[f"A{left_row}"].font = bold
        ws[f"A{left_row}"].alignment = left
        ws[f"B{left_row}"] = value
        ws[f"B{left_row}"].alignment = left
        left_row += 1

    for label, value in info_rows_middle:
        ws[f"C{middle_row}"] = label
        ws[f"C{middle_row}"].font = bold
        ws[f"C{middle_row}"].alignment = left
        ws[f"D{middle_row}"] = value
        ws[f"D{middle_row}"].alignment = left
        middle_row += 1
    
    # Write right info (cols D–E)
    for label, value in info_rows_right:
        ws[f"E{right_row}"] = label
        ws[f"E{right_row}"].font = bold
        ws[f"E{right_row}"].alignment = left
        ws[f"F{right_row}"] = value
        ws[f"F{right_row}"].alignment = left
        right_row += 1

    # Leave a blank line
    row = max(left_row, middle_row, right_row) + 1

    # ---- Table Header ----
    ws["A" + str(row)] = "Description"
    ws["B" + str(row)] = "Responsible Person"
    ws["C" + str(row)] = "Process"

    # Merge D–F for "Comments"
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    ws["D" + str(row)] = "Comments"

    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.font = bold
        cell.alignment = center
        cell.border = thin_border

    row += 1
    
    # ---- Sample table data (use your actual comments here) ----
    sections = [
        ("Material Concerns (SMS)", "W/House/QC/CS",
            ["Fabric issue", "Trims issue", "Bonding / Seam Tape parameter"]),
        ("Logo Application", "QC/Prdn",
            ["Printing issue", "Embroidery issue", "Heat transfer issue"]),
        ("Construction", "APM/QC/TECH", [""]),
        ("Stitching and Sewing (Special tools, M/C, Needle etc.)", "APM/QC/TECH/Maint", [""]),
        ("SMS Notes & Risk Analysis", "QC & QA/CS", [""]),
        ("Size Set Evaluation & Comments", "QC", [""]), 
        ("PP meeting Special attention (SMS review, size set review, Pattern review, risk point review & Customer comments review)", 
            "W/House/QC/CS/APM/TECH/Mechanic/QA/PATTERN/CUTTING", [""]), 
        ("Bulk Production", "APM/QC/TECH", 
            ["Line layout", "Machine Layout", "Special Tools(Gage, Guide, Folder)", "KPM, KPI & Self Inspection"]), 
        ("1st Output", "Individual Team", ["QC comments", "QA comments", "TQS comments"]),
        ("Factory Inline issue", "APM & QC", ["Factory Feedback & Solution"]),
        ("Customer Comments", "QA", ["Inline comments", "Final Comments"]), 
        ("Others", "", [""])
    ]

    for section_name, responsible, processes in sections:
        first_row = row
        for process in processes:
            ws.cell(row=row, column=3).value = process
            ws.cell(row=row, column=3).alignment = left
            ws.cell(row=row, column=3).border = thin_border

            # Comments spanning D–F
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
            comment_obj = Comment.objects.filter(style=style, process=process).first()
            ws.cell(row=row, column=4).value = comment_obj.comment_text if comment_obj else ""
            ws.cell(row=row, column=4).alignment = left

            # Apply border to merged area
            for col in range(4, 7):
                ws.cell(row=row, column=col).border = thin_border

            row += 1


        # ---- Set custom row heights for specific sections ----
        if section_name == "Stitching and Sewing (Special tools, M/C, Needle etc.)":
            ws.row_dimensions[first_row].height = 53
        elif section_name == "Size Set Evaluation & Comments":
            ws.row_dimensions[first_row].height = 33
        elif section_name.startswith("PP meeting Special attention"):
            ws.row_dimensions[first_row].height = 100

        # Merge and fill section name & responsible person
        ws.merge_cells(start_row=first_row, start_column=1, end_row=row - 1, end_column=1)
        ws.cell(row=first_row, column=1).value = section_name
        ws.cell(row=first_row, column=1).alignment = center
        set_border(ws, first_row, 1, row - 1, 1, thin_border)

        ws.merge_cells(start_row=first_row, start_column=2, end_row=row - 1, end_column=2)
        ws.cell(row=first_row, column=2).value = responsible
        ws.cell(row=first_row, column=2).alignment = center
        set_border(ws, first_row, 2, row - 1, 2, thin_border)

    # ---- Return Excel ----
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"Style_{style.style_no}_Summary.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response